"""
Read weld geometry and load cases from Excel (.xlsx) for EuroWeldGroup.

Expected workbook structure
----------------------------
Sheet **Settings** (columns A:B, header row 1)
    Parameter | Value
    fu, gamma_M2, steel_grade, beta_w, weld_size_z, throat_a,
    reduce_ends, min_throat_mm, check_min_length, thicker_part_mm,
    search_z_min, search_z_max (optional bounds for required-size search)

    Provide either weld_size_z or throat_a (not both). Leave beta_w empty to use steel_grade.

Sheet **WeldGeometry** (header row 1)
    label | y1_mm | z1_mm | y2_mm | z2_mm

Sheet **Loadcases** (header row 1)
    name | Fx_N | Fy_N | Fz_N |
    Pfx_x_mm | Pfx_y_mm | Pfx_z_mm |
    Pfy_x_mm | Pfy_y_mm | Pfy_z_mm |
    Pfz_x_mm | Pfz_y_mm | Pfz_z_mm

Units match main.py: N, mm, MPa for fu.
"""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from main import (
    MultiLoadCaseResult,
    WeldSegment,
    check_multiple_loadcases,
    print_multiple_loadcases_result,
    print_weld_group_result,
    required_weld_size_for_multiple_loadcases,
    weld_group_check_from_component_points,
)


def _truthy(cell_val: Any) -> bool:
    if cell_val is None:
        return False
    if isinstance(cell_val, bool):
        return cell_val
    s = str(cell_val).strip().lower()
    return s in ("1", "true", "yes", "y", "ja")


def _float(cell_val: Any, default: Optional[float] = None) -> Optional[float]:
    if cell_val is None or (isinstance(cell_val, str) and not str(cell_val).strip()):
        return default
    return float(cell_val)


def _str_or_none(cell_val: Any) -> Optional[str]:
    if cell_val is None:
        return None
    s = str(cell_val).strip()
    return s if s else None


def read_settings_sheet(ws) -> Dict[str, Any]:
    """Parse Settings sheet: column A = parameter name, B = value (row 1 = header)."""
    params: Dict[str, Any] = {}
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0] is None:
            continue
        key = str(row[0]).strip().lower().replace(" ", "_")
        if not key:
            continue
        params[key] = row[1]

    required = ["fu"]
    for k in required:
        if k not in params or params[k] is None or str(params[k]).strip() == "":
            raise ValueError(f"Settings sheet: missing required parameter '{k}'.")

    fu = float(params["fu"])
    gm = _float(params.get("gamma_m2"), 1.25)
    if gm is None:
        gm = 1.25
    gamma_M2 = float(gm)

    steel_grade = _str_or_none(params.get("steel_grade"))
    beta_w = _float(params.get("beta_w"), None)

    wz = _float(params.get("weld_size_z"), None)
    ta = _float(params.get("throat_a"), None)
    if wz is not None and ta is not None:
        raise ValueError("Settings: provide only one of weld_size_z or throat_a.")
    if wz is None and ta is None:
        raise ValueError("Settings: provide weld_size_z or throat_a.")

    reduce_ends = _truthy(params.get("reduce_ends", True))
    mtm = _float(params.get("min_throat_mm"), 3.0)
    min_throat_mm = float(mtm) if mtm is not None else 3.0
    check_min_length = _truthy(params.get("check_min_length", True))
    thicker = _float(params.get("thicker_part_mm"), None)

    sz_min = _float(params.get("search_z_min"), 3.0)
    sz_max = _float(params.get("search_z_max"), 25.0)
    search_z_min = float(sz_min) if sz_min is not None else 3.0
    search_z_max = float(sz_max) if sz_max is not None else 25.0

    return {
        "fu": fu,
        "gamma_M2": gamma_M2,
        "steel_grade": steel_grade,
        "beta_w": beta_w,
        "weld_size_z": wz,
        "throat_a": ta,
        "reduce_ends": reduce_ends,
        "min_throat_mm": min_throat_mm,
        "check_min_length": check_min_length,
        "thicker_part_mm": thicker,
        "search_z_min": search_z_min,
        "search_z_max": search_z_max,
    }


def _norm_header(h: Any) -> str:
    if h is None:
        return ""
    s = str(h).strip().lower()
    s = re.sub(r"\s+", "_", s)
    s = s.replace("(mm)", "mm").replace("(n)", "n")
    return s


def read_weld_geometry_sheet(ws) -> List[WeldSegment]:
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        raise ValueError("WeldGeometry sheet is empty.")
    headers = [_norm_header(c) for c in rows[0]]

    def col(*names: str) -> int:
        for n in names:
            if n in headers:
                return headers.index(n)
        raise ValueError(
            f"WeldGeometry: missing column (need one of {names!r}). Found: {headers}"
        )

    i_label = col("label", "segment", "name", "id")
    i_y1 = col("y1_mm", "y1")
    i_z1 = col("z1_mm", "z1")
    i_y2 = col("y2_mm", "y2")
    i_z2 = col("z2_mm", "z2")

    segments: List[WeldSegment] = []
    for r in rows[1:]:
        if r is None or all(v is None or str(v).strip() == "" for v in r):
            continue
        y1 = float(r[i_y1])
        z1 = float(r[i_z1])
        y2 = float(r[i_y2])
        z2 = float(r[i_z2])
        lab = r[i_label] if i_label < len(r) else ""
        label = str(lab).strip() if lab is not None else ""
        segments.append(WeldSegment(y1, z1, y2, z2, label))
    if not segments:
        raise ValueError("WeldGeometry: no data rows found.")
    return segments


def read_loadcases_sheet(ws) -> List[Dict[str, Any]]:
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        raise ValueError("Loadcases sheet is empty.")
    headers = [_norm_header(c) for c in rows[0]]

    def idx(name: str) -> int:
        if name not in headers:
            raise ValueError(f"Loadcases: missing column '{name}'. Found: {headers}")
        return headers.index(name)

    i_name = idx("name")
    i_fx = idx("fx_n")
    i_fy = idx("fy_n")
    i_fz = idx("fz_n")
    for c in (
        "pfx_x_mm",
        "pfx_y_mm",
        "pfx_z_mm",
        "pfy_x_mm",
        "pfy_y_mm",
        "pfy_z_mm",
        "pfz_x_mm",
        "pfz_y_mm",
        "pfz_z_mm",
    ):
        idx(c)

    def tri(r: Tuple, px: str, py: str, pz: str) -> Tuple[float, float, float]:
        return (float(r[idx(px)]), float(r[idx(py)]), float(r[idx(pz)]))

    loadcases: List[Dict[str, Any]] = []
    for r in rows[1:]:
        if r is None or all(v is None or str(v).strip() == "" for v in r):
            continue
        nm = r[i_name]
        name = str(nm).strip() if nm is not None else ""
        if not name:
            name = f"LC{len(loadcases) + 1}"
        loadcases.append(
            {
                "name": name,
                "Fx": float(r[i_fx]),
                "Fy": float(r[i_fy]),
                "Fz": float(r[i_fz]),
                "Pfx": tri(r, "pfx_x_mm", "pfx_y_mm", "pfx_z_mm"),
                "Pfy": tri(r, "pfy_x_mm", "pfy_y_mm", "pfy_z_mm"),
                "Pfz": tri(r, "pfz_x_mm", "pfz_y_mm", "pfz_z_mm"),
            }
        )
    if not loadcases:
        raise ValueError("Loadcases: no data rows found.")
    return loadcases


@dataclass
class ExcelModelInput:
    settings: Dict[str, Any]
    segments: List[WeldSegment]
    loadcases: List[Dict[str, Any]]


@dataclass
class ExcelAnalysis:
    """Results of running the checker on a workbook (single computation pass)."""

    model: ExcelModelInput
    check_kw: Dict[str, Any]
    multi_result: MultiLoadCaseResult
    required_size: Optional[Dict[str, Any]]


def read_weld_workbook(path: str | Path) -> ExcelModelInput:
    from openpyxl import load_workbook

    p = Path(path)
    if not p.is_file():
        raise FileNotFoundError(p)

    wb = load_workbook(p, data_only=True)
    try:
        if "Settings" not in wb.sheetnames:
            raise ValueError(f"Workbook must contain sheet 'Settings'. Got: {wb.sheetnames}")
        if "WeldGeometry" not in wb.sheetnames:
            raise ValueError(
                f"Workbook must contain sheet 'WeldGeometry'. Got: {wb.sheetnames}"
            )
        if "Loadcases" not in wb.sheetnames:
            raise ValueError(f"Workbook must contain sheet 'Loadcases'. Got: {wb.sheetnames}")

        settings = read_settings_sheet(wb["Settings"])
        segments = read_weld_geometry_sheet(wb["WeldGeometry"])
        loadcases = read_loadcases_sheet(wb["Loadcases"])
    finally:
        wb.close()

    return ExcelModelInput(settings=settings, segments=segments, loadcases=loadcases)


def _kwargs_for_check(settings: Dict[str, Any]) -> Dict[str, Any]:
    kw: Dict[str, Any] = {
        "fu": settings["fu"],
        "gamma_M2": settings["gamma_M2"],
        "steel_grade": settings["steel_grade"],
        "beta_w": settings["beta_w"],
        "reduce_ends": settings["reduce_ends"],
        "min_throat_mm": settings["min_throat_mm"],
        "check_min_length": settings["check_min_length"],
        "thicker_part_mm": settings["thicker_part_mm"],
    }
    if settings["weld_size_z"] is not None:
        kw["weld_size_z"] = settings["weld_size_z"]
    else:
        kw["throat_a"] = settings["throat_a"]
    return kw


def analyze_excel_workbook(path: str | Path, *, required_size: bool = True) -> ExcelAnalysis:
    """Load workbook, run all load cases, and optionally required-weld-size search."""
    data = read_weld_workbook(path)
    kw = _kwargs_for_check(data.settings)
    multi = check_multiple_loadcases(
        segments=data.segments, loadcases=data.loadcases, **kw
    )
    req: Optional[Dict[str, Any]] = None
    if required_size:
        z_min = float(data.settings["search_z_min"])
        z_max = float(data.settings["search_z_max"])
        req_kw = {k: v for k, v in kw.items() if k not in ("weld_size_z", "throat_a")}
        try:
            req = required_weld_size_for_multiple_loadcases(
                segments=data.segments,
                loadcases=data.loadcases,
                z_min=z_min,
                z_max=z_max,
                **req_kw,
            )
        except ValueError as e:
            req = {"_error": str(e)}
    return ExcelAnalysis(model=data, check_kw=kw, multi_result=multi, required_size=req)


def print_excel_analysis(a: ExcelAnalysis, *, required_size: bool = True) -> None:
    """Print text reports for a completed workbook analysis."""
    data = a.model
    kw = a.check_kw

    first = data.loadcases[0]
    r0 = weld_group_check_from_component_points(
        segments=data.segments,
        Fx=first["Fx"],
        Fy=first["Fy"],
        Fz=first["Fz"],
        Pfx=first["Pfx"],
        Pfy=first["Pfy"],
        Pfz=first["Pfz"],
        **kw,
    )
    print(f"\n--- First load case: {first['name']} ---")
    print_weld_group_result(r0)

    print_multiple_loadcases_result(a.multi_result)

    if required_size and a.required_size is not None:
        if "_error" in a.required_size:
            print(f"\nRequired weld size search skipped: {a.required_size['_error']}")
        else:
            print("\nREQUIRED WELD SIZE FOR ALL LOADCASES")
            print("=" * 100)
            for k, v in a.required_size.items():
                if isinstance(v, float):
                    print(f"{k:<28} = {v:.4f}")
                else:
                    print(f"{k:<28} = {v}")


def run_excel_check(path: str | Path, *, required_size: bool = True) -> None:
    """Load workbook, run multi-loadcase check, print reports; optionally required weld size."""
    a = analyze_excel_workbook(path, required_size=required_size)
    print_excel_analysis(a, required_size=required_size)


def write_sample_workbook(path: str | Path) -> None:
    """Create a documented sample .xlsx matching read_weld_workbook expectations."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # remove default sheet; we'll add ordered sheets
    default = wb.active
    wb.remove(default)

    ws_about = wb.create_sheet("About", 0)
    ws_about["A1"] = "EuroWeldGroup — Excel input format"
    ws_about["A1"].font = Font(bold=True, size=14)
    ws_about["A3"] = (
        "Edit Settings, WeldGeometry, and Loadcases. Units: N, mm, MPa (fu). "
        "Weld segments lie in the y-z plane (x = 0 at the weld plane)."
    )
    ws_about["A5"] = "Run:  python weld_excel.py sample_weld_input.xlsx"

    ws_s = wb.create_sheet("Settings")
    ws_s.append(["Parameter", "Value"])
    settings_rows = [
        ("fu", 510.0),
        ("gamma_M2", 1.25),
        ("steel_grade", "S355"),
        ("beta_w", ""),
        ("weld_size_z", 6.0),
        ("throat_a", ""),
        ("reduce_ends", True),
        ("min_throat_mm", 3.0),
        ("check_min_length", True),
        ("thicker_part_mm", 16.0),
        ("search_z_min", 3.0),
        ("search_z_max", 16.0),
    ]
    for row in settings_rows:
        ws_s.append(list(row))

    ws_w = wb.create_sheet("WeldGeometry")
    ws_w.append(["label", "y1_mm", "z1_mm", "y2_mm", "z2_mm"])
    b, h = 200.0, 100.0
    for lab, y1, z1, y2, z2 in [
        ("bot", -b / 2, -h / 2, b / 2, -h / 2),
        ("rhs", b / 2, -h / 2, b / 2, h / 2),
        ("top", b / 2, h / 2, -b / 2, h / 2),
        ("lhs", -b / 2, h / 2, -b / 2, -h / 2),
    ]:
        ws_w.append([lab, y1, z1, y2, z2])

    ws_l = wb.create_sheet("Loadcases")
    hdr = [
        "name",
        "Fx_N",
        "Fy_N",
        "Fz_N",
        "Pfx_x_mm",
        "Pfx_y_mm",
        "Pfx_z_mm",
        "Pfy_x_mm",
        "Pfy_y_mm",
        "Pfy_z_mm",
        "Pfz_x_mm",
        "Pfz_y_mm",
        "Pfz_z_mm",
    ]
    ws_l.append(hdr)
    sample_cases = [
        (
            "LC1",
            25000,
            40000,
            15000,
            120,
            0,
            0,
            120,
            0,
            35,
            120,
            25,
            0,
        ),
        (
            "LC2",
            18000,
            10000,
            30000,
            120,
            0,
            0,
            120,
            0,
            35,
            120,
            25,
            0,
        ),
        (
            "LC3",
            -10000,
            25000,
            5000,
            120,
            0,
            0,
            120,
            0,
            35,
            120,
            25,
            0,
        ),
    ]
    for row in sample_cases:
        ws_l.append(list(row))

    for ws in (ws_s, ws_w, ws_l):
        for i, _ in enumerate(ws[1], start=1):
            ws.column_dimensions[get_column_letter(i)].width = 14

    out = Path(path)
    wb.save(out)


def main() -> None:
    parser = argparse.ArgumentParser(description="Run EuroWeldGroup from an Excel workbook.")
    parser.add_argument(
        "workbook",
        nargs="?",
        default="sample_weld_input.xlsx",
        help="Path to .xlsx (default: sample_weld_input.xlsx)",
    )
    parser.add_argument(
        "--no-required-size",
        action="store_true",
        help="Skip required weld size search",
    )
    parser.add_argument(
        "--create-sample",
        metavar="OUT.xlsx",
        help="Write a sample workbook to OUT.xlsx and exit",
    )
    parser.add_argument(
        "--pdf",
        metavar="OUT.pdf",
        help="Write sizing report PDF (inputs, results, coordinate system, weld plots)",
    )
    args = parser.parse_args()
    if args.create_sample:
        write_sample_workbook(args.create_sample)
        print(f"Sample workbook written to {args.create_sample}")
        return
    a = analyze_excel_workbook(args.workbook, required_size=not args.no_required_size)
    if args.pdf:
        from weld_report_pdf import write_sizing_pdf_from_analysis

        write_sizing_pdf_from_analysis(a, args.pdf)
        print(f"PDF written to {args.pdf}")
    print_excel_analysis(a, required_size=not args.no_required_size)


if __name__ == "__main__":
    main()
